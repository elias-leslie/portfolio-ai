"""CMS QHP landscape PUF ingest for the retirement ACA estimator (item D part c).

Downloads the "QHP Landscape Individual Medical" public use file from
data.healthcare.gov, filters to the household's counties, and stores the
age-rated premiums by metal tier in ``aca_marketplace_plans``. The
estimator anchors on the age-21 premium: Florida uses the federal
default age curve, so any age rates as ``premium_age_21 x curve factor``
(the sample-age columns let that be verified against published rates).

The download URL pattern was verified against the live data.healthcare.gov
metastore for PY2026 (dataset 6fe7fb77-7291-4104-952f-7c7e2c5d0c45,
distribution ``datafile/py2026/individual_market_medical.zip``). CMS
publishes the next plan year's file around open enrollment (November).
"""

from __future__ import annotations

import tempfile
import zipfile
from datetime import UTC, datetime
from importlib import import_module
from pathlib import Path
from typing import Any

import httpx

from app.logging_config import get_logger

logger = get_logger(__name__)

LANDSCAPE_URL_TEMPLATE = (
    "https://data.healthcare.gov/datafile/py{plan_year}/individual_market_medical.zip"
)
# (state_code, FIPS county code) pairs to keep. Pinellas FL is the
# household's home county; widen via the counties parameter if they move.
DEFAULT_COUNTIES: tuple[tuple[str, str], ...] = (("FL", "12103"),)
DEFAULT_PLAN_YEAR = 2026
_DOWNLOAD_TIMEOUT_SECONDS = 300.0

# Normalized PUF header -> stored field. Headers carry stray trailing
# whitespace in the published file ("Premium Adult Individual Age 30 "),
# so lookups strip both sides.
_COLUMN_FIELDS = {
    "State Code": "state_code",
    "FIPS County Code": "fips_county_code",
    "County Name": "county_name",
    "Metal Level": "metal_level",
    "Issuer Name": "issuer_name",
    "Plan ID (Standard Component)": "plan_id",
    "Plan Marketing Name": "plan_marketing_name",
    "Standardized Plan Option": "standardized_plan_option",
    "Plan Type": "plan_type",
    "Rating Area": "rating_area",
    "EHB Percent of Total Premium": "ehb_percent",
    "Premium Child Age 0-14": "premium_child_age_0_14",
    "Premium Child Age 18": "premium_child_age_18",
    "Premium Adult Individual Age 21": "premium_age_21",
    "Premium Adult Individual Age 27": "premium_age_27",
    "Premium Adult Individual Age 30": "premium_age_30",
    "Premium Adult Individual Age 40": "premium_age_40",
    "Premium Adult Individual Age 50": "premium_age_50",
    "Premium Adult Individual Age 60": "premium_age_60",
    "Medical Deductible - Individual - Standard": "medical_deductible_individual",
    "Medical Maximum Out Of Pocket - Individual - Standard": "medical_moop_individual",
}
_TEXT_FIELDS = {
    "state_code",
    "county_name",
    "metal_level",
    "issuer_name",
    "plan_id",
    "plan_marketing_name",
    "standardized_plan_option",
    "plan_type",
    "rating_area",
}
_MONEY_FIELDS = {
    "premium_child_age_0_14",
    "premium_child_age_18",
    "premium_age_21",
    "premium_age_27",
    "premium_age_30",
    "premium_age_40",
    "premium_age_50",
    "premium_age_60",
    "medical_deductible_individual",
    "medical_moop_individual",
}


def parse_money(value: Any) -> float | None:
    """'$1,510.31' -> 1510.31; blanks and non-numeric markers -> None."""
    if value is None:
        return None
    if isinstance(value, int | float):
        return round(float(value), 2)
    text = str(value).strip().replace("$", "").replace(",", "")
    if not text:
        return None
    try:
        return round(float(text), 2)
    except ValueError:
        return None


def parse_percent(value: Any) -> float | None:
    """'100.00%' -> 100.0."""
    if value is None:
        return None
    if isinstance(value, int | float):
        return float(value)
    text = str(value).strip().rstrip("%")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def normalize_fips(value: Any) -> str:
    """County FIPS arrives as text, int, or float; store 5-digit text."""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    return text.zfill(5) if text.isdigit() else text


def parse_landscape_xlsx(
    xlsx_path: Path,
    *,
    counties: tuple[tuple[str, str], ...],
) -> list[dict[str, Any]]:
    """Extract the configured counties' plans from the PUF workbook.

    The sheet has a banner row above the real header, so the header is
    located by scanning for the row whose first cell is 'State Code'.
    Columns are resolved by name to survive ordering drift between plan
    years; a missing required column raises rather than mis-mapping.
    """
    import openpyxl  # noqa: PLC0415 — heavyweight; only needed at ingest time

    wanted = {(state, fips) for state, fips in counties}
    workbook = openpyxl.load_workbook(xlsx_path, read_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        rows = sheet.iter_rows(values_only=True)
        column_index: dict[str, int] | None = None
        for row in rows:
            if row and str(row[0] or "").strip() == "State Code":
                column_index = {
                    str(cell).strip(): idx for idx, cell in enumerate(row) if cell is not None
                }
                break
        if column_index is None:
            raise ValueError("PUF header row not found (no 'State Code' column)")
        missing = [name for name in _COLUMN_FIELDS if name not in column_index]
        if missing:
            raise ValueError(f"PUF columns missing: {missing}")

        plans: list[dict[str, Any]] = []
        for row in rows:
            state = str(row[column_index["State Code"]] or "").strip()
            fips = normalize_fips(row[column_index["FIPS County Code"]])
            if (state, fips) not in wanted:
                continue
            record: dict[str, Any] = {"state_code": state, "fips_county_code": fips}
            for column, field in _COLUMN_FIELDS.items():
                if field in ("state_code", "fips_county_code"):
                    continue
                value = row[column_index[column]]
                if field in _MONEY_FIELDS:
                    record[field] = parse_money(value)
                elif field == "ehb_percent":
                    record[field] = parse_percent(value)
                else:
                    record[field] = str(value).strip() if value is not None else None
            plans.append(record)
        return plans
    finally:
        workbook.close()


def _extract_xlsx(zip_path: Path, target_dir: Path) -> Path:
    with zipfile.ZipFile(zip_path) as archive:
        names = [n for n in archive.namelist() if n.lower().endswith(".xlsx")]
        if not names:
            raise ValueError(f"No .xlsx inside {zip_path.name}")
        archive.extract(names[0], target_dir)
    return target_dir / names[0]


class AcaMarketplaceIngestService:
    """Download, parse, and persist the landscape PUF for tracked counties."""

    def __init__(self, storage: Any | None = None) -> None:
        self.storage = storage or import_module("app.storage").get_storage()

    def ingest(
        self,
        *,
        plan_year: int = DEFAULT_PLAN_YEAR,
        counties: tuple[tuple[str, str], ...] = DEFAULT_COUNTIES,
        xlsx_path: str | None = None,
        url: str | None = None,
    ) -> dict[str, Any]:
        source_url = url or LANDSCAPE_URL_TEMPLATE.format(plan_year=plan_year)
        if xlsx_path is not None:
            plans = parse_landscape_xlsx(Path(xlsx_path), counties=counties)
            source_url = f"file://{xlsx_path}"
        else:
            with tempfile.TemporaryDirectory(prefix="aca-puf-") as tmp:
                tmp_dir = Path(tmp)
                zip_path = tmp_dir / "landscape.zip"
                logger.info("aca_puf_download_start", url=source_url)
                with httpx.stream(
                    "GET", source_url, timeout=_DOWNLOAD_TIMEOUT_SECONDS, follow_redirects=True
                ) as response:
                    response.raise_for_status()
                    with zip_path.open("wb") as handle:
                        for chunk in response.iter_bytes():
                            handle.write(chunk)
                plans = parse_landscape_xlsx(
                    _extract_xlsx(zip_path, tmp_dir), counties=counties
                )

        by_metal: dict[str, int] = {}
        for plan in plans:
            metal = str(plan.get("metal_level") or "unknown")
            by_metal[metal] = by_metal.get(metal, 0) + 1

        now = datetime.now(UTC)
        fips_list = [fips for _, fips in counties]
        with self.storage.connection() as conn:
            conn.execute(
                "DELETE FROM aca_marketplace_plans WHERE plan_year = %s AND fips_county_code = ANY(%s)",
                [plan_year, fips_list],
            )
            for plan in plans:
                conn.execute(
                    """
                    INSERT INTO aca_marketplace_plans (
                        plan_year, state_code, fips_county_code, county_name,
                        metal_level, issuer_name, plan_id, plan_marketing_name,
                        standardized_plan_option, plan_type, rating_area,
                        ehb_percent, premium_child_age_0_14, premium_child_age_18,
                        premium_age_21, premium_age_27, premium_age_30,
                        premium_age_40, premium_age_50, premium_age_60,
                        medical_deductible_individual, medical_moop_individual,
                        source_url, ingested_at
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                              %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    [
                        plan_year,
                        plan["state_code"],
                        plan["fips_county_code"],
                        plan.get("county_name"),
                        plan.get("metal_level"),
                        plan.get("issuer_name"),
                        plan.get("plan_id"),
                        plan.get("plan_marketing_name"),
                        plan.get("standardized_plan_option"),
                        plan.get("plan_type"),
                        plan.get("rating_area"),
                        plan.get("ehb_percent"),
                        plan.get("premium_child_age_0_14"),
                        plan.get("premium_child_age_18"),
                        plan.get("premium_age_21"),
                        plan.get("premium_age_27"),
                        plan.get("premium_age_30"),
                        plan.get("premium_age_40"),
                        plan.get("premium_age_50"),
                        plan.get("premium_age_60"),
                        plan.get("medical_deductible_individual"),
                        plan.get("medical_moop_individual"),
                        source_url,
                        now,
                    ],
                )
            conn.commit()

        summary = {
            "plan_year": plan_year,
            "counties": [f"{state}-{fips}" for state, fips in counties],
            "inserted": len(plans),
            "by_metal": by_metal,
            "source_url": source_url,
        }
        logger.info("aca_puf_ingested", **summary)
        return summary

    def list_plans(
        self,
        *,
        plan_year: int | None = None,
        fips_county_code: str | None = None,
    ) -> list[dict[str, Any]]:
        where = ["TRUE"]
        params: list[Any] = []
        if plan_year is not None:
            where.append("plan_year = %s")
            params.append(plan_year)
        if fips_county_code is not None:
            where.append("fips_county_code = %s")
            params.append(fips_county_code)
        sql = f"""
            SELECT plan_year, state_code, fips_county_code, county_name,
                   metal_level, issuer_name, plan_id, plan_marketing_name,
                   standardized_plan_option, plan_type, rating_area, ehb_percent,
                   premium_child_age_0_14, premium_child_age_18, premium_age_21,
                   premium_age_27, premium_age_30, premium_age_40, premium_age_50,
                   premium_age_60, medical_deductible_individual,
                   medical_moop_individual, source_url, ingested_at
            FROM aca_marketplace_plans
            WHERE {" AND ".join(where)}
            ORDER BY plan_year DESC, metal_level, premium_age_21 ASC NULLS LAST
        """
        with self.storage.connection() as conn:
            rows = conn.execute(sql, params).fetchall()
        columns = [
            "plan_year", "state_code", "fips_county_code", "county_name",
            "metal_level", "issuer_name", "plan_id", "plan_marketing_name",
            "standardized_plan_option", "plan_type", "rating_area", "ehb_percent",
            "premium_child_age_0_14", "premium_child_age_18", "premium_age_21",
            "premium_age_27", "premium_age_30", "premium_age_40", "premium_age_50",
            "premium_age_60", "medical_deductible_individual",
            "medical_moop_individual", "source_url", "ingested_at",
        ]
        plans: list[dict[str, Any]] = []
        for row in rows:
            record = dict(zip(columns, row, strict=True))
            for field in _MONEY_FIELDS | {"ehb_percent"}:
                if record[field] is not None:
                    record[field] = float(record[field])
            if record["ingested_at"] is not None:
                record["ingested_at"] = record["ingested_at"].isoformat()
            plans.append(record)
        return plans
