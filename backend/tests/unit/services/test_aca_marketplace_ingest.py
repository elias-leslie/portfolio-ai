"""Unit tests for the CMS QHP landscape PUF parser (retirement item D-c)."""

from __future__ import annotations

from pathlib import Path

import pytest

from app.services.aca_marketplace_ingest_service import (
    normalize_fips,
    parse_landscape_xlsx,
    parse_money,
    parse_percent,
)

_HEADER = [
    "State Code",
    "FIPS County Code",
    "County Name",
    "Metal Level",
    "Issuer Name",
    "Plan ID (Standard Component)",
    "Plan Marketing Name",
    "Standardized Plan Option",
    "Plan Type",
    "Rating Area",
    "EHB Percent of Total Premium",
    "Premium Child Age 0-14",
    "Premium Child Age 18",
    "Premium Adult Individual Age 21",
    "Premium Adult Individual Age 27",
    # Trailing whitespace as published in the real PY2026 file.
    "Premium Adult Individual Age 30 ",
    "Premium Adult Individual Age 40 ",
    "Premium Adult Individual Age 50 ",
    "Premium Adult Individual Age 60 ",
    "Medical Deductible - Individual - Standard",
    "Medical Maximum Out Of Pocket - Individual - Standard",
]


def _write_fixture(path: Path, rows: list[list[object]]) -> Path:
    import openpyxl

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Individual_Market_Medical"
    sheet.append(["=SUBTOTAL(3,A3:A1048576)", "displayed records"])  # banner row
    sheet.append(_HEADER)
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    return path


def _plan_row(*, state: str = "FL", fips: object = 12103, metal: str = "Silver") -> list[object]:
    return [
        state,
        fips,
        "Pinellas",
        metal,
        "Florida Blue",
        "16842FL0120033",
        "BlueCare Silver 2026",
        "Not Applicable",
        "EPO",
        "Rating Area 52",
        "100.00%",
        "$425.10",
        "$556.49",
        "$766.78",
        "$804.21",
        "$870.30",
        "$979.94",
        "$1,369.47",
        "$2,081.04",
        "$5,000",
        "$9,100",
    ]


def test_parse_money_and_percent_and_fips() -> None:
    assert parse_money("$1,510.31") == 1510.31
    assert parse_money(556.49) == 556.49
    assert parse_money("") is None
    assert parse_money("Not Applicable") is None
    assert parse_percent("100.00%") == 100.0
    assert parse_percent(None) is None
    assert normalize_fips(12103) == "12103"
    assert normalize_fips(12103.0) == "12103"
    assert normalize_fips("8013") == "08013"


def test_parse_filters_to_requested_counties(tmp_path: Path) -> None:
    fixture = _write_fixture(
        tmp_path / "puf.xlsx",
        [
            _plan_row(),
            _plan_row(fips=12057, metal="Bronze"),  # Hillsborough — filtered out
            _plan_row(state="GA", fips=13121),
        ],
    )

    plans = parse_landscape_xlsx(fixture, counties=(("FL", "12103"),))

    assert len(plans) == 1
    plan = plans[0]
    assert plan["state_code"] == "FL"
    assert plan["fips_county_code"] == "12103"
    assert plan["metal_level"] == "Silver"
    assert plan["plan_id"] == "16842FL0120033"
    assert plan["premium_age_21"] == 766.78
    assert plan["premium_age_60"] == 2081.04
    assert plan["premium_child_age_0_14"] == 425.10
    assert plan["ehb_percent"] == 100.0
    assert plan["medical_moop_individual"] == 9100.0


def test_parse_raises_on_missing_required_column(tmp_path: Path) -> None:
    import openpyxl

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["State Code", "FIPS County Code", "County Name"])
    path = tmp_path / "bad.xlsx"
    workbook.save(path)

    with pytest.raises(ValueError, match="PUF columns missing"):
        parse_landscape_xlsx(path, counties=(("FL", "12103"),))


def test_parse_raises_when_header_row_absent(tmp_path: Path) -> None:
    import openpyxl

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["banner", "only"])
    path = tmp_path / "headerless.xlsx"
    workbook.save(path)

    with pytest.raises(ValueError, match="header row not found"):
        parse_landscape_xlsx(path, counties=(("FL", "12103"),))
